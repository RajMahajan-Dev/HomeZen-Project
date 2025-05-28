from flask import render_template, request, redirect, url_for, flash, session, Blueprint ,Response, send_file
from flask_login import LoginManager, login_user, login_required, logout_user, current_user, UserMixin
from datetime import datetime as dt
from models import User, Proffessional, Client, Service, Order,Review
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import  secure_filename
from sqlalchemy import func
import os
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

from app import db


auth = Blueprint('auth', __name__)
routes = Blueprint('routes', __name__)

@auth.route('/', methods=['POST', 'GET'])
def Home():
    return render_template('home.html')

@auth.route('/login', methods=['POST', 'GET'])
def login():
    if request.method == 'POST':
        user_handle = request.form.get('user_handle')
        user_password = request.form.get('user_password')
        user = User.query.filter_by(user_handle=user_handle).first()
        if user:
            if check_password_hash(user.user_password, user_password):
                if user.user_role == 1:
                    client = Client.query.filter_by(client_user_id=user.user_id).first()
                    if client.client_flaged == True:
                        flash('Your account has been Suspended. Contact Admin', 'danger')
                        return render_template('login.html')
                    else:
                        login_user(user)
                        flash('Logined Successfully', 'success')
                        return redirect(url_for('routes.dashboard'))
                elif user.user_role == 2:
                    proffessional = Proffessional.query.filter_by(proffessional_user_id=user.user_id).first()
                    if proffessional.proffessional_flaged == True:
                        flash('Your account has been Suspended. Contact Admin', 'danger')
                        return render_template('login.html')
                    else:
                        login_user(user)
                        flash('Logined Successfully', 'success')
                        return redirect(url_for('routes.dashboard'))  
                elif user.user_role == 3:
                    login_user(user)
                    flash('Logined Successfully', 'success')
                    return redirect(url_for('routes.dashboard')) 
            else:
                flash('Password is incorrect' , "danger")
        else:
            flash('User does not exist. Create an account', "warning")
            return render_template('register.html')

    return render_template('login.html')

@auth.route('/register', methods=['POST', 'GET'])
def register():
    user = User.query.filter_by(user_role=3).first()
    print(user)
    if not user:
        new_user = User(user_name='Admin', user_handle='admin', user_password=generate_password_hash('admin1122',method='pbkdf2:sha256',salt_length=8),user_email='admin@contact.us',user_role=3)
        db.session.add(new_user)
        db.session.commit()

    if request.method == 'POST':
        user_name = request.form.get('user_name')
        user_handle = request.form.get('user_handle')
        user_password = request.form.get('user_password')
        user_role = int(request.form.get('user_role'))
        user_email = request.form.get('user_email')

        user = User.query.filter_by(user_handle=user_handle).first()
        if user:
            flash('User already exists. Login Now' ,"danger")
        else:
            new_user = User(user_name=user_name, user_handle=user_handle, user_password=generate_password_hash(user_password,method='pbkdf2:sha256',salt_length=8),user_email=user_email ,user_role=user_role)
            db.session.add(new_user)
            db.session.commit()
            if new_user.user_role == 2:
                new_proffessional = Proffessional(proffessional_user_id=new_user.user_id, proffessional_company_name='None', proffessional_email=new_user.user_email, proffessional_description='None', proffessional_Service_type='None', proffessional_years_of_experience=0, proffessional_location='None')
                db.session.add(new_proffessional)
                db.session.commit()
            elif new_user.user_role == 1:
                new_client = Client(client_user_id=new_user.user_id, client_bio='None', client_email=new_user.user_email, client_location='None', client_total_spend=0)
                db.session.add(new_client)
                db.session.commit()
            flash('Registration Successful. Login Now' , "success")
            return redirect(url_for('auth.login'))
    return render_template('register.html')

@auth.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.login'))

@routes.route('/dashboard', methods=['GET','POST'])
@login_required
def dashboard():
    Client_this = Client.query.filter_by(client_user_id=current_user.user_id).first()
    if current_user.user_role == 1:
        if request.method == 'POST':
            if request.form.get('form_type') == "Completed":
                order_id = request.form.get('order_id')
                order = Order.query.filter_by(order_id=order_id).first()
                if order:
                    order.order_status = "Completed"
                    Client_this.client_total_spend += order.order_payment_amount
                    order.order_date_completed = dt.now()
                    db.session.commit()
                    flash('Order Completed successfully', 'success')
                else:
                    flash('Order not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "Review":
                order_id = request.form.get('order_id')
                order = Order.query.filter_by(order_id=order_id).first()
                comment = request.form.get('comment')
                if comment == "":
                    comment = "---"
                if order:
                    new_review = Review(
                        review_order_id=order.order_id,
                        review_rating=request.form.get('rating'),
                        review_comment=comment
                    )
                    db.session.add(new_review)
                    db.session.commit()
                    flash('Review added successfully', 'success')
                else:
                    flash('Order not found', 'danger')
                return redirect(url_for('routes.dashboard'))
        
        Ranks = db.session.query(
                Client,
                Client.client_user_id,
                Client.client_total_spend,
                User.user_name
            ).join(
                User,
                Client.client_user_id == User.user_id
            ).order_by(Client.client_total_spend.desc()).all()
        Rank_No = 0
        for Rank in Ranks:
            Rank_No+=1
            if Rank.client_user_id == current_user.user_id:
                break


        Order_this = Order.query.filter_by(order_client_id=Client_this.client_user_id).all()
        ordered = 0
        count = 0
        orders = db.session.query(
                Order,
                Order.order_id,
                Order.order_name,
                Order.order_payment_amount,
                Order.order_date_created,
                Order.order_date_accepted,
                Order.order_date_completed,
                Order.order_status,
                Order.order_flaged,
                Review.review_rating,
                Review.review_comment,
                Proffessional.proffessional_company_name
            ).join(
                Proffessional,
                Order.order_proffessional_id == Proffessional.proffessional_user_id
            ).outerjoin(
                Review,
                Order.order_id == Review.review_order_id
            ).filter(
                Order.order_client_id == current_user.user_id
            ).order_by(Order.order_id.desc()).all()
            
        if Order_this:
            ordered =1
            for ord in Order_this:
                count += 1
        
        return render_template('client_dashboard.html', client=Client_this ,order=Order_this, ordered=ordered,count=count, orders_all=orders,Rank_No=Rank_No)
    
    
    elif current_user.user_role == 2:
        
        if request.method == 'POST':
            if request.form.get('form_type') == "update_order_status":
                order_id = request.form.get('order_id')
                order = Order.query.filter_by(order_id=order_id).first()
                if order:
                    order.order_status = request.form.get('order_status')
                    order.order_date_accepted = dt.now()
                    db.session.commit()
                    flash('Order '+ order.order_status + ' successfully', 'success')
                else:
                    flash('Order not found', 'danger')
                return redirect(url_for('routes.dashboard'))
           
            if request.form.get('form_type') == "Completed":
                order_id = request.form.get('order_id')
                order = Order.query.filter_by(order_id=order_id).first()
                Client_this = Client.query.filter_by(client_user_id=order.order_client_id).first()
                if order:
                    order.order_status = "Completed"
                    Client_this.client_total_spend += order.order_payment_amount
                    order.order_date_completed = dt.now()
                    db.session.commit()
                    flash('Order Completed successfully', 'success')
                else:
                    flash('Order not found', 'danger')
                return redirect(url_for('routes.dashboard'))

        if request.method == 'POST':
            if request.form.get('form_type') == "Get_Excel":
                ord = db.session.query(
                    Order,
                    Order.order_id,
                    Order.order_name,
                    Order.order_payment_amount,
                    Order.order_date_created,
                    Order.order_date_accepted,
                    Order.order_date_completed,
                    Order.order_status,
                    Order.order_flaged,
                    Review.review_rating,
                    Review.review_comment,
                    User.user_name
                ).join(
                    User,
                    Order.order_client_id == User.user_id
                ).outerjoin(
                    Review,
                    Order.order_id == Review.review_order_id
                ).filter(
                    Order.order_proffessional_id == current_user.user_id
                ).order_by(Order.order_date_created.desc())

                info = db.session.query(
                    Proffessional,
                    Proffessional.proffessional_user_id,
                    Proffessional.proffessional_company_name,
                    Proffessional.proffessional_Service_type,
                    Proffessional.proffessional_location,
                    Proffessional.proffessional_pincode,
                    Proffessional.proffessional_rating,
                ).filter(
                    Proffessional.proffessional_user_id == current_user.user_id
                ).first()

                professional_info = {
                    "User ID": info.proffessional_user_id,
                    "Company Name": info.proffessional_company_name,
                    "Service Type": info.proffessional_Service_type,
                    "Location": info.proffessional_location,
                    "Pincode": f"{info.proffessional_pincode}",
                    "Rating": f"{info.proffessional_rating:.2f}"
                }

                df_professional = pd.DataFrame([professional_info])

                order_data = [
                    {
                        "Order ID": r.order_id,
                        "Order Name": r.order_name,
                        "Payment Amount": r.order_payment_amount,
                        "Client Name": r.user_name,
                        "Order Status": r.order_status,
                        "Flagged": r.order_flaged,
                        "Review Rating": r.review_rating if r.review_rating else "None",
                        "Review Comment": r.review_comment if r.review_comment else "None",
                        "Created Date": r.order_date_created.strftime('%Y-%m-%d'),
                        "Accepted Date": r.order_date_accepted.strftime('%Y-%m-%d') if r.order_date_accepted else "Not Accepted",
                        "Completed Date": r.order_date_completed.strftime('%Y-%m-%d') if r.order_date_completed else "Not Completed"
                    }
                    for r in ord
                ]

                df_orders = pd.DataFrame(order_data)

                wb = Workbook()
                ws = wb.active
                ws.title = "Professional and Orders Data"

                ws.append(["Professional Information"])
                ws.append([])
                for row in dataframe_to_rows(df_professional, index=False, header=True):
                    ws.append(row)

                ws.append([])
                ws.append([])
                ws.append(["Orders Data"])
                ws.append([])

                for row in dataframe_to_rows(df_orders, index=False, header=True):
                    ws.append(row)
                for cell in ws[1]:
                    cell.font = Font(bold=True, size=14)
                for cell in ws[3]:
                    cell.font = Font(bold=True, size=11)
                for cell in ws[9]:
                    cell.font = Font(bold=True, size=11)
                for cell in ws[7]:
                    cell.font = Font(bold=True, size=14)

                output = BytesIO()
                wb.save(output)
                wb.save(f"excel_output/professional_orders_data_{info.proffessional_user_id}.xlsx")
                output.seek(0)
                return send_file(output,
                                 mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                 as_attachment=True,
                                 download_name=f'professional_orders_data_{info.proffessional_user_id}.xlsx')



        Proffessional_this = Proffessional.query.filter_by(proffessional_user_id=current_user.user_id).first()
        Order_this = db.session.query(
                        Order,
                        Order.order_id,
                        Order.order_name,
                        Order.order_payment_amount,
                        Order.order_date_created,
                        Order.order_date_accepted,
                        Order.order_date_completed,
                        Order.order_status,
                        Order.order_flaged,
                        Review.review_rating,
                        Review.review_comment,
                        User.user_name
                    ).join(
                        User,
                        Order.order_client_id == User.user_id
                    ).outerjoin(
                        Review,
                        Order.order_id == Review.review_order_id
                    ).filter(
                        Order.order_proffessional_id == current_user.user_id
                    ).all()
        flag = True
        for ord in Order_this:
            if ord.order_status == "Accepted":
                flag = False 
                break
        if flag == False:
            Proffessional_this.proffessional_available = False
            db.session.commit()
        else:
            Proffessional_this.proffessional_available = True
            db.session.commit()
            
        rating = db.session.query(
            Review.review_rating,
            func.count(Review.review_id).label('count')
        ).join(
            Order,
            Review.review_order_id == Order.order_id
        ).join(
            Service,
            Order.order_service_id == Service.service_id
        ).join(
            Proffessional,
            Service.service_proffessional_id == Proffessional.proffessional_user_id
        ).filter(
            Proffessional.proffessional_user_id == current_user.user_id,
            Review.review_flaged == False
        ).group_by(
            Review.review_rating
        ).all()
        Sum =0
        c=0
        rating_dict = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        Average = 0
        for rate, count in rating:
            rating_dict[int(rate//1)] = count
            Sum += rate * count
            c += count
        if Sum != 0:
            Average = Sum/c
        proffessional = Proffessional.query.filter_by(proffessional_user_id=current_user.user_id).first()
        if Average != 0:
            proffessional.proffessional_rating = Average
            Average = 0
            db.session.commit()
        
        return render_template('proffessional_dashboard.html', proffessional=Proffessional_this,orders=Order_this)
        

    elif current_user.user_role == 3:
        if request.method == 'POST':
            if request.form.get('form_type') == "flag_client":
                client_id = request.form.get('client_id')
                client = Client.query.filter_by(client_id=client_id).first()
                if client:
                    if request.form.get('flag_status') == "True":
                        client.client_flaged = True
                        flash('Client flagged successfully', 'success')
                    else:
                        client.client_flaged = False    
                        flash('Client Unflagged successfully', 'success')                
                    db.session.commit()
                else:
                    flash('Client not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "flag_proffessional":
                proffessional_id = request.form.get('proffessional_id')
                proffessional = Proffessional.query.filter_by(proffessional_id=proffessional_id).first()
                if proffessional:
                    if request.form.get('flag_status') == "True":
                        proffessional.proffessional_flaged = True
                        flash('Proffessional flagged successfully', 'success')
                    else:
                        proffessional.proffessional_flaged = False
                        flash('Proffessional Unflagged successfully', 'success')                    
                    db.session.commit()
                else:
                    flash('Proffessional not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "flag_service":
                service_id = request.form.get('service_id')
                service = Service.query.filter_by(service_id=service_id).first()
                if service:
                    if request.form.get('flag_status') == "True":
                        service.service_flaged = True
                        flash('Service flagged successfully', 'success')
                    else:
                        service.service_flaged = False
                        flash('Service Unflagged successfully', 'success')                    
                    db.session.commit()
                else:
                    flash('Service not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "flag_order":
                order_id = request.form.get('order_id')
                order = Order.query.filter_by(order_id=order_id).first()
                if order:
                    if request.form.get('flag_status') == "True":
                        order.order_flaged = True
                        flash('Order flagged successfully', 'success')
                    else:
                        order.order_flaged = False
                        flash('Order Unflagged successfully', 'success')                   
                    db.session.commit()
                else:
                    flash('Order not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "flag_review":
                review_id = request.form.get('review_id')
                review = Review.query.filter_by(review_id=review_id).first()
                if review:
                    review.review_flaged = True
                    db.session.commit()
                    flash('Review flagged successfully', 'success')
                else:
                    flash('Review not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "delete_client":
                client_id = request.form.get('client_id')
                client = Client.query.filter_by(client_user_id=client_id).first()
                user = User.query.filter_by(user_id=client_id).first()
                if client:
                    db.session.delete(client)
                    db.session.delete(user)
                    db.session.commit()
                    flash('Client deleted successfully', 'success')
                else:
                    flash('Client not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "delete_proffessional":
                proffessional_id = request.form.get('proffessional_id')
                proffessional = Proffessional.query.filter_by(proffessional_id=proffessional_id).first()
                user = User.query.filter_by(user_id=proffessional.proffessional_user_id).first()
                if proffessional:
                    db.session.delete(proffessional)
                    db.session.delete(user)
                    db.session.commit()
                    flash('Proffessional deleted successfully', 'success')
                else:
                    flash('Proffessional not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "delete_service":
                service_id = request.form.get('service_id')
                service = Service.query.filter_by(service_id=service_id).first()
                if service:
                    db.session.delete(service)
                    db.session.commit()
                    flash('Service deleted successfully', 'success')
                else:
                    flash('Service not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "delete_order":
                order_id = request.form.get('order_id')
                order = Order.query.filter_by(order_id=order_id).first()
                review = Review.query.filter_by(review_order_id=order_id).first()
                if order:
                    db.session.delete(order)
                    if review:
                        db.session.delete(review)
                    db.session.commit()
                    flash('Order deleted successfully', 'success')
                else:
                    flash('Order not found', 'danger')
                return redirect(url_for('routes.dashboard'))
            elif request.form.get('form_type') == "delete_review":
                review_id = request.form.get('review_id')
                review = Review.query.filter_by(review_id=review_id).first()
                if review:
                    db.session.delete(review)
                    db.session.commit()
                    flash('Review deleted successfully', 'success')
                else:
                    flash('Review not found', 'danger')
                return redirect(url_for('routes.dashboard'))


        Client_this = Client.query.all()
        Proffessional_this = Proffessional.query.all()

        Service_this = db.session.query(
            Service,
            Service.service_id,
            Service.service_name,
            Service.service_description,
            Service.service_price,
            Service.service_photo,
            Service.service_flaged,
            Service.service_date_created,
            Service.service_proffessional_id,

            Proffessional,
            Proffessional.proffessional_company_name
        ).join(
            Proffessional,
            Service.service_proffessional_id == Proffessional.proffessional_user_id
        ).all()

        order_this = db.session.query(
            Order,
            Order.order_id,
            Order.order_name,
            Order.order_payment_amount,
            Order.order_status,
            Order.order_flaged,
            Proffessional,
            Proffessional.proffessional_company_name,
            Review,
            Review.review_rating,
            Review.review_comment,
            User,
            User.user_name
        ).join(
            Proffessional,
            Order.order_proffessional_id == Proffessional.proffessional_user_id
        ).join(
            User,
            Order.order_client_id == User.user_id
        ).outerjoin(
            Review,
            Order.order_id == Review.review_order_id
        ).all()

        review_this = db.session.query(
            Review,
            Review.review_id,
            Review.review_rating,
            Review.review_comment,
            Order,
            Order.order_name,
            User,
            User.user_name,
            Proffessional,
            Proffessional.proffessional_company_name
        ).join(
            Order,
            Review.review_order_id == Order.order_id
        ).join(
            User,
            Order.order_client_id == User.user_id
        ).join(
            Proffessional,
            Order.order_proffessional_id == Proffessional.proffessional_user_id
        ).all()


        return render_template('admin_dashboard.html', clients=Client_this, proffessionals=Proffessional_this ,services=Service_this, orders=order_this, reviews=review_this)
        

    else:
        return render_template('Error.html')


@routes.route('/search', methods=['GET','POST'])
@login_required
def search():
    if current_user.user_role == 1:
        if request.method == 'POST':
            S_type= request.form.get('search_type')
            S_query = request.form.get('search_query')

            try:
                services = db.session.query(
                    Service,
                    Service.service_id,
                    Service.service_name,
                    Service.service_description,
                    Service.service_price,
                    Service.service_photo,
                    Proffessional.proffessional_Service_type,
                    Proffessional.proffessional_company_name
                ).join(
                    Proffessional,
                    Service.service_proffessional_id == Proffessional.proffessional_user_id
                ).filter(Service.service_flaged == False)
                if S_type == "service_name":
                    print("S0" , S_query)
                    service = services.filter(Service.service_name.ilike(f"%{S_query}%")).all()
                    if service:
                        print("S1" , service)
                        services = services.filter(Service.service_name.ilike(f"%{S_query}%"))
                    else:
                        services = services.filter(Proffessional.proffessional_Service_type.ilike(f"%{S_query}%"))
                    

                elif S_type == "location":
                    services = services.filter(Proffessional.proffessional_location.ilike(f"%{S_query}%"))
                elif S_type == "pincode":
                    services = services.filter(Proffessional.proffessional_pincode.ilike(f"%{S_query}%"))
                
                services = services.all()

                if not services:
                    flash('No services found,Try different Search', 'warning')
                return render_template('client_search.html', services=services)
                
            except Exception as e:
                flash('Error searching services', 'danger')
                return render_template('client_search.html')
        return render_template('client_search.html')
    elif current_user.user_role == 2:
        if request.method == 'POST':
            S_type= request.form.get('search_type')
            S_query = request.form.get('search_query')


            try:
                services = db.session.query(
                    Service,
                    Service.service_id,
                    Service.service_name,
                    Service.service_description,
                    Service.service_price,
                    Service.service_photo,
                    Service.service_flaged,
                    Proffessional.proffessional_Service_type,
                    Proffessional.proffessional_company_name
                ).join(
                    Proffessional,
                    Service.service_proffessional_id == Proffessional.proffessional_user_id
                )
                if S_type == "service_name":
                    print("S0" , S_query)
                    service = services.filter(Service.service_name.ilike(f"%{S_query}%")).all()
                    if service:
                        print("S1" , service)
                        services = services.filter(Service.service_name.ilike(f"%{S_query}%"))
                    else:
                        services = services.filter(Proffessional.proffessional_Service_type.ilike(f"%{S_query}%"))
                    

                elif S_type == "location":
                    services = services.filter(Proffessional.proffessional_location.ilike(f"%{S_query}%"))
                elif S_type == "pincode":
                    services = services.filter(Proffessional.proffessional_pincode.ilike(f"%{S_query}%"))
                
                services = services.all()

                if not services:
                    flash('No services found,Try different Search', 'warning')
                return render_template('proffessional_search.html', services=services)
                
            except Exception as e:
                flash('Error searching services', 'danger')
                return render_template('proffessional_search.html')
        return render_template('proffessional_search.html')
    elif current_user.user_role == 3:
        if request.method == 'POST':
            S_type= request.form.get('search_type')
            S_query = request.form.get('search_query')

            try:
                Proffessionals = None
                if S_type == "location":
                    Proffessionals = Proffessional.query.filter(Proffessional.proffessional_location.ilike(f"%{S_query}%")).all()
                elif S_type == "pincode":
                    Proffessionals = Proffessional.query.filter(Proffessional.proffessional_pincode.ilike(f"%{S_query}%")).all()
                elif S_type == "proffessional_company_name":
                    Proffessionals = Proffessional.query.filter(Proffessional.proffessional_company_name.ilike(f"%{S_query}%")).all()
                elif S_type == "service_type":
                    Proffessionals = Proffessional.query.filter(Proffessional.proffessional_Service_type.ilike(f"%{S_query}%")).all()

                if not Proffessionals:
                    flash('No services found,Try different Search', 'warning')
                return render_template('admin_search.html',proffessionals=Proffessionals)
                
            except Exception as e:
                flash('Error searching services', 'danger')
                return render_template('admin_search.html')
        return render_template('admin_search.html')
    
@routes.route('/profile', methods=['GET','POST'])
@login_required
def profile():
    if current_user.user_role == 2:
        if request.method == 'POST':
            try:
                proffessional = Proffessional.query.filter_by(proffessional_user_id=current_user.user_id).first()
                proffessional.proffessional_email = request.form.get('Pemail')
                proffessional.proffessional_location = request.form.get('Plocation')
                proffessional.proffessional_description = request.form.get('Pdescription')
                proffessional.proffessional_years_of_experience = request.form.get('Pyoe')
                proffessional.proffessional_company_name = request.form.get('Pcompany_name')
                proffessional.proffessional_Service_type = request.form.get('PService_type')
                proffessional.proffessional_pincode = request.form.get('Ppincode')
                U_folder = 'static/Proff_Data'
                photo= None
                if 'Pphoto' in request.files:
                    file = request.files["Pphoto"]
                    if file:
                        filename= secure_filename(file.filename)
                        uni_file = f"{dt.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
                        file_path = os.path.join(U_folder, uni_file)
                        file.save(file_path)
                        photo=f"static/Proff_Data/{uni_file}"
                        proffessional.proffessional_photo = photo
                db.session.commit()
                flash('Profile updated successfully', 'success')
            except Exception as e:
                db.session.rollback()
                flash('Error updating profile', 'danger')  
            return redirect(url_for('routes.profile'))


        proffessional = Proffessional.query.filter_by(proffessional_user_id=current_user.user_id).first()
        return render_template('proffessional_profile.html' , proffessional=proffessional)
    elif current_user.user_role == 1:
        if request.method == 'POST':
            try:
                client = Client.query.filter_by(client_user_id=current_user.user_id).first()
                client.client_email = request.form.get('Cemail')
                client.client_location = request.form.get('Clocation')
                client.client_bio = request.form.get('Cbio')
                client.client_pincode = request.form.get('Cpincode')
                U_folder = 'static/Client_Data'
                photo= None
                if 'Cphoto' in request.files:
                    file = request.files["Cphoto"]
                    if file:
                        filename= secure_filename(file.filename)
                        uni_file = f"{dt.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
                        file_path = os.path.join(U_folder, uni_file)
                        file.save(file_path)
                        photo=f"static/Client_Data/{uni_file}"
                        client.client_photo = photo
                db.session.commit()
                flash('Profile updated successfully', 'success')
            except Exception as e:
                db.session.rollback()
                flash('Error updating profile', 'danger')  
            return redirect(url_for('routes.profile'))

        client = Client.query.filter_by(client_user_id=current_user.user_id).first()
        return render_template('client_profile.html' , client=client)
    elif current_user.user_role == 3:
        User_this = User.query.filter_by(user_id=current_user.user_id).first()
        return render_template('admin_profile.html',user=User_this)
    else:
        return render_template('Error.html')
    
    
@routes.route('/services', methods=['GET','POST'])
@login_required
def services():
    
    if current_user.user_role == 2:

        if request.method== "POST":
            if request.form.get('form_type') == "add":
                Sname = request.form.get('Sname')
                Sdescription = request.form.get('Sdescription')
                Sprice = request.form.get('Sprice')

                U_folder = 'static/Proff_Data'
                photo= None
                if 'Sphoto' in request.files:
                    file = request.files["Sphoto"]
                    if file:
                        filename= secure_filename(file.filename)
                        uni_file = f"{dt.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
                        file_path = os.path.join(U_folder, uni_file)
                        file.save(file_path)
                        photo=f"static/Proff_Data/{uni_file}"

                new_service = Service(
                    service_name=Sname,
                    service_description=Sdescription,
                    service_price=Sprice,
                    service_photo=photo,
                    service_proffessional_id = current_user.user_id
                )
                db.session.add(new_service)
                db.session.commit()
                flash('Service added successfully', 'success')
                return redirect(url_for('routes.services'))
            elif request.form.get('form_type') == "edit":
                    try:
                        service_id = request.form.get('Sno')
                        service = Service.query.filter_by(service_id=service_id).first()
                        if service:
                            service.service_name = request.form.get('Sname')
                            service.service_description = request.form.get('Sdescription')
                            service.service_price = request.form.get('Sprice')
                            U_folder = 'static/Proff_Data'
                            if 'Sphoto' in request.files:
                                file = request.files["Sphoto"]
                                if file:
                                    filename= secure_filename(file.filename)
                                    uni_file = f"{dt.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
                                    file_path = os.path.join(U_folder, uni_file)
                                    file.save(file_path)
                                    service.service_photo=f"static/Proff_Data/{uni_file}"

                            db.session.commit()
                            flash('Service updated successfully', 'success')
                        else:
                            flash('Service not found', 'danger')

                    except Exception as e:
                        db.session.rollback()
                        flash('Error updating service', 'danger')

                    return redirect(url_for('routes.services'))
        

            elif request.form.get('form_type') == "delete":
                try:
                    service_id = request.form.get('service_id')
                    service = Service.query.filter_by(service_id=service_id).first()
                    if service:
                        db.session.delete(service)
                        db.session.commit()
                        flash('Service deleted successfully', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash('Error deleting service', 'danger')
                return redirect(url_for('routes.services'))



        proff = Proffessional.query.filter_by(proffessional_user_id=current_user.user_id).first()
        services = Service.query.filter_by(service_proffessional_id=proff.proffessional_user_id).all()
        return render_template('proffessional_services.html' , services=services)
    elif current_user.user_role == 1:  #Client
        if request.method== "POST":
            if request.form.get('form_type') == "buy":
                try:
                    service_id = request.form.get('service_id')
                    service = Service.query.filter_by(service_id=service_id).first()
                    if service:
                        new_order = Order(
                            order_service_id=service.service_id,
                            order_client_id=current_user.user_id,
                            order_proffessional_id=service.service_proffessional_id,
                            order_payment_amount=service.service_price,
                            order_name=service.service_name
                        )
                        db.session.add(new_order)
                        db.session.commit()
                        flash('Service ordered successfully', 'success')
                    else:
                        flash('Service not found', 'danger')
                except Exception as e:
                    db.session.rollback()
                    flash('Error ordering service', 'danger')
                if request.form.get('from') == "Search":
                    print("C1")
                    return redirect(url_for('routes.search'))
                elif request.form.get('from') == "services":
                    print("C2")
                    return redirect(url_for('routes.services'))

        
        services = db.session.query(
                Service,
                Service.service_id,
                Service.service_name,
                Service.service_description,
                Service.service_price,
                Service.service_photo,
                Proffessional.proffessional_rating,
                Proffessional.proffessional_Service_type,
                Proffessional.proffessional_company_name
            ).join(
                Proffessional,
                Service.service_proffessional_id == Proffessional.proffessional_user_id
            ).filter(
                Service.service_flaged == False
            ).all()
        S_types = db.session.query(
            Proffessional.proffessional_Service_type
        ).distinct(
            Proffessional.proffessional_Service_type
        ).all()
        return render_template('client_services.html',services=services, S_types=S_types)
    elif current_user.user_role == 3:
        return render_template('admin_services.html')
    else:
        return render_template('Error.html')

    
@routes.route('/our_growth')
def our_growth():
    proff_count = len(Proffessional.query.all())
    client_count = len(Client.query.all())
    js_data6 = [['Users','Preffessional', 'Clients']]
    js_data6.append(['Users', proff_count, client_count])

    service_counts = db.session.query(
        Proffessional.proffessional_Service_type,
        func.count(Service.service_id).label('count')
    ).join(
        Service,
        Service.service_proffessional_id == Proffessional.proffessional_user_id
    ).filter(
        Service.service_flaged == False
    ).group_by(
        Proffessional.proffessional_Service_type
    ).all()

    counts_dict = {type: count for type, count in service_counts}
    js_data = [['Service', 'Available']]
    js_data.extend([[key, value] for key, value in counts_dict.items()])

    orders =db.session.query(
        Proffessional.proffessional_Service_type,
        func.count(Order.order_id).label('count')
    ).join(
        Service,
        Order.order_service_id == Service.service_id
    ).join(
        Proffessional,
        Service.service_proffessional_id == Proffessional.proffessional_user_id
    ).filter(
        Order.order_flaged == False
    ).group_by(
        Proffessional.proffessional_Service_type
    ).all()
    js_data5 = [['Service', 'Orders']]
    for Service_type, count in orders:
        js_data5.append([Service_type, count])

    return render_template('our_growth.html' , proff_count=proff_count, client_count=client_count, js_data=js_data,js_data5=js_data5,js_data6=js_data6)

@routes.route('/stats' , methods=['GET','POST'])
@login_required
def stats():

    service_counts = db.session.query(
        Proffessional.proffessional_Service_type,
        func.count(Service.service_id).label('count')
    ).join(
        Service,
        Service.service_proffessional_id == Proffessional.proffessional_user_id
    ).filter(
        Service.service_flaged == False
    ).group_by(
        Proffessional.proffessional_Service_type
    ).all()


    # Convert to dictionary for easy access
    counts_dict = {type: count for type, count in service_counts}
    js_data = [['Service', 'Available']]  # Header row
    js_data.extend([[key, value] for key, value in counts_dict.items()])

    order_stats = db.session.query(
        Proffessional.proffessional_Service_type,
        Order.order_status,
        func.count(Order.order_id).label('count')
    ).join(
        Service,
        Order.order_service_id == Service.service_id
    ).join(
        Proffessional,
        Service.service_proffessional_id == Proffessional.proffessional_user_id
    ).filter(
        Order.order_flaged == False
    ).group_by(
        Proffessional.proffessional_Service_type,
        Order.order_status
    ).all()
    js_data2 = [['Service', 'Completed', 'Rejected', 'Accepted']]

    # Group counts by service type
    service_stats = {}
    for type, status, count in order_stats:
        if type not in service_stats:
            service_stats[type] = {'Completed': 0, 'Rejected': 0, 'Accepted': 0}
        service_stats[type][status] = count

    # Add data rows
    for service_type, counts in service_stats.items():
        js_data2.append([
            service_type,
            counts['Completed'],
            counts['Rejected'],
            counts['Accepted']
        ])
    if current_user.user_role == 2:
        # Get all reviews for the proffessional

        rating = db.session.query(
            Review.review_rating,
            func.count(Review.review_id).label('count')
        ).join(
            Order,
            Review.review_order_id == Order.order_id
        ).join(
            Service,
            Order.order_service_id == Service.service_id
        ).join(
            Proffessional,
            Service.service_proffessional_id == Proffessional.proffessional_user_id
        ).filter(
            Proffessional.proffessional_user_id == current_user.user_id,
            Review.review_flaged == False
        ).group_by(
            Review.review_rating
        ).all()
        Sum =0
        c=0
        rating_dict = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        Average = 0
        for rate, count in rating:
            rating_dict[int(rate//1)] = count
            Sum += rate * count
            c += count
        if Sum != 0:
            Average = Sum/c

        return render_template('P_summary.html' ,js_data=js_data, js_data2=js_data2, rating=rating_dict, Average=round(Average,2))
    elif current_user.user_role == 1:

        return render_template('C_summary.html' ,js_data=js_data, js_data2=js_data2)
    elif current_user.user_role == 3:
        proff_count = len(Proffessional.query.all())
        client_count = len(Client.query.all())
        js_data6 = [['Users','Preffessional', 'Clients']]
        js_data6.append(['Users', proff_count, client_count])
    
        service_counts = db.session.query(
            Proffessional.proffessional_Service_type,
            func.count(Service.service_id).label('count')
        ).join(
            Service,
            Service.service_proffessional_id == Proffessional.proffessional_user_id
        ).filter(
            Service.service_flaged == False
        ).group_by(
            Proffessional.proffessional_Service_type
        ).all()
    
        counts_dict = {type: count for type, count in service_counts}
        js_data = [['Service', 'Available']]
        js_data.extend([[key, value] for key, value in counts_dict.items()])
    
        orders =db.session.query(
            Proffessional.proffessional_Service_type,
            func.count(Order.order_id).label('count'),
            func.count(Review.review_id).label('count')
        ).join(
            Service,
            Order.order_service_id == Service.service_id
        ).join(
            Proffessional,
            Service.service_proffessional_id == Proffessional.proffessional_user_id
        ).outerjoin(
            Review,
            Order.order_id == Review.review_order_id
        ).filter(
            Order.order_flaged == False
        ).group_by(
            Proffessional.proffessional_Service_type
        ).order_by(
            func.count(Order.order_id)
        ).all()
        js_data5 = [['Service', 'Orders', 'Reviews']]
        for Service_type, count,rcount in orders:
            js_data5.append([Service_type, count,rcount])
        
        proff_flage_count = len(Proffessional.query.filter_by(proffessional_flaged=True).all())
        client_flage_count = len(Client.query.filter_by(client_flaged=True).all())
        client_count = len(Client.query.all())
        proff_count = len(Proffessional.query.all())
        service_count = len(Service.query.all())
        order_count = len(Order.query.all())
        review_count = db.session.query(
            Review.review_rating,
            func.count(Review.review_id).label('count')
        ).group_by(
            Review.review_rating
        ).all()
        Service_flage_count = len(Service.query.filter_by(service_flaged=True).all())
        order_flage_count = len(Order.query.filter_by(order_flaged=True).all())

        js_data7 = [['User', 'Count']]
        js_data7.append(['Proffessional', proff_count])
        js_data7.append(['Client', client_count])
        js_data7.append(['Proffessional Flaged', proff_flage_count])
        js_data7.append(['Client Flaged', client_flage_count])

        js_data8 = [['Service', 'Count']]
        js_data8.append(['Service', service_count])
        js_data8.append(['Service Flaged', Service_flage_count])

        js_data9 = [['Order', 'Count']]
        js_data9.append(['Order', order_count])
        js_data9.append(['Order Flaged', order_flage_count])

        js_data10 = [['Rating', 'Count']]
        for rating , count in review_count:
            js_data10.append([rating//1, count])

        return render_template('A_summary.html' , proff_count=proff_count, client_count=client_count,js_data2=js_data2, js_data=js_data,js_data5=js_data5,js_data6=js_data6 , js_data7=js_data7, js_data8=js_data8, js_data9=js_data9, js_data10=js_data10)
    else:
        return render_template('Error.html')


@routes.route('/<path:path>')
def catch_all(path):
    return render_template('Error.html')